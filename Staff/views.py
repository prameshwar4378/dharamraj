from django.shortcuts import render,redirect,get_object_or_404
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .forms import *
from Developer.models import *
from .filters import *
from django.contrib import messages
from django.http import Http404
# Create your views here.
from django.db.models import Sum, F, ExpressionWrapper, fields
from datetime import date,timedelta
from django.db.models.functions import Coalesce
from django.contrib.auth.decorators import login_required

from django.contrib.auth.decorators import login_required, user_passes_test

def is_staff(user):
    return user.is_authenticated and hasattr(user, 'is_staff') and user.is_staff

def staff_required(view_func):
    decorated_view_func = login_required(user_passes_test(is_staff, login_url='/accounts/login')(view_func))
    return decorated_view_func

 
@staff_required
def dashboard(request):
    # Get today's date
    today_date = date.today()
    end_of_day = today_date + timedelta(days=1)
    
    total_grand_total_today = Invoice.objects.filter(invoice_date=today_date).aggregate(Sum('grand_total'))['grand_total__sum'] or 0
    total_gst_amount_today = Invoice.objects.filter(invoice_date=today_date).aggregate(Sum('total_gst_amount'))['total_gst_amount__sum'] or 0
    total_quantity_today = InvoiceItem.objects.filter(invoice__invoice_date=today_date).aggregate(Sum('quantity'))['quantity__sum'] or 0
    total_invoices_today = Invoice.objects.filter(invoice_date=today_date).count()

    total_credited_amount_today = Account.objects.filter(
        is_credit=True, 
        transaction_date__range=(today_date, end_of_day)
    ).aggregate(Sum('amount'))['amount__sum'] or 0

    total_debited_amount_today = Account.objects.filter(
        is_credit=False, 
        transaction_date__range=(today_date, end_of_day)
    ).aggregate(Sum('amount'))['amount__sum'] or 0

    print(total_credited_amount_today) 

    total_credited_amount = Account.objects.filter(is_credit=True).aggregate(Sum('amount'))['amount__sum'] or 0
    total_debited_amount = Account.objects.filter(is_credit=False).aggregate(Sum('amount'))['amount__sum'] or 0
    total_balance = Account.objects.aggregate(
        total_balance=Sum(ExpressionWrapper(
            F('amount') * F('is_credit') - F('amount') * ~F('is_credit'),
            output_field=fields.FloatField()
        ))
    )['total_balance'] or 0
 
    start_date = date.today() - timedelta(days=30)
    top_five_dealers = Invoice.objects.filter(invoice_date__gte=start_date).values('dealer__business_name').annotate(total_amount=Coalesce(Sum('grand_total'), 0)).order_by('-total_amount')[:5]
    top_five_dealer_total_amount = [item['total_amount'] for item in top_five_dealers]
    top_five_dealer_name = [item['dealer__business_name'] for item in top_five_dealers]

    # Calculate top five sale products in the last 30 days
    top_five_sale_products = InvoiceItem.objects.filter(invoice__invoice_date__gte=start_date).values('product__product_name').annotate(total_quantity=Coalesce(Sum('quantity'), 0)).order_by('-total_quantity')[:5]
    top_five_sale_product_quantity = [item['total_quantity'] for item in top_five_sale_products]
    top_five_sale_product_labels = [item['product__product_name'] for item in top_five_sale_products]
    
    context = {
        'total_grand_total_today': total_grand_total_today,
        'total_gst_amount_today': total_gst_amount_today,
        'total_quantity_today': total_quantity_today,
        'total_invoices_today': total_invoices_today,
        'total_credited_amount_today': total_credited_amount_today,
        'total_debited_amount_today': total_debited_amount_today,
        # 'total_balance_today': total_balance_today,
        'top_five_dealer_total_amount': top_five_dealer_total_amount,
        'top_five_dealer_name': top_five_dealer_name,
        'top_five_sale_product_quantity': top_five_sale_product_quantity,
        'top_five_sale_product_labels': top_five_sale_product_labels,
        'total_credited_amount': total_credited_amount,
        'total_debited_amount': total_debited_amount,
        'total_balance': total_balance,
    }

    return render(request, 'staff__dashboard.html', context)


 
@staff_required
def dealer_list(request):
    try:
        dealer_rec = Dealer.objects.select_related().order_by('dealer_name')

        # Add pagination
        page = request.GET.get('page', 1)
        paginator = Paginator(dealer_rec, 25)  # Show 10 dealers per page
        try:
            dealer_rec = paginator.page(page)
        except PageNotAnInteger:
            dealer_rec = paginator.page(1)
        except EmptyPage:
            dealer_rec = paginator.page(paginator.num_pages)

        create_dealer_fm = DealerCreationForm()
        update_dealer_fm = DealerUpdateForm()
        context = {
            'dealer_rec': dealer_rec,
            'create_dealer_fm': create_dealer_fm,
            'update_dealer_fm': update_dealer_fm
        }
        return render(request, 'staff__dealer_list.html', context)
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

@staff_required
def create_dealer(request):
    try:
        form = DealerCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Dealer Created Successfully.....')
        else:
            error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
            messages.warning(request, error_message)
        return redirect('/staff/dealer_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

@staff_required
def update_dealer(request): 
    try:
        if request.method == 'POST':
            dealer_id = request.POST.get('txt_id') 
            dealer_instance = get_object_or_404(Dealer, id=dealer_id)
            form = DealerUpdateForm(request.POST, instance=dealer_instance)
            if form.is_valid():
                form.save()
                messages.success(request, 'Dealer Updated Successfully.')
                return redirect('/staff/dealer_list/')
            else:
                error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
                messages.warning(request, error_message)
            return redirect('/staff/dealer_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

@staff_required
def delete_dealer(request, id):
    try:
        dealer = get_object_or_404(Dealer, id=id)
        dealer.delete()
        messages.success(request, 'Dealer Deleted Successfully.')
        return redirect('/staff/dealer_list/')
    except Http404:
        return render(request, '404.html')


@staff_required
def product_list(request):
    try:
        product_rec = Product.objects.select_related()
        create_product_fm = ProductCreationForm()
        update_product_fm = ProductUpdateForm()
        add_stock = PurchaseCreationForm()
        update_stock = PurchaseUpdateForm()
        context = {'product_rec': product_rec,'add_stock':add_stock,'update_stock':update_stock, 'create_product_fm': create_product_fm, 'update_product_fm': update_product_fm}
        return render(request, 'staff__product_list.html', context)
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

@staff_required
def create_product(request):
    try:
        form = ProductCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product Created Successfully.....')
        else:
            error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
            messages.warning(request, error_message)
        return redirect('/staff/product_list/')
    except Exception as e:
        # Log the exception if needed
        messages.error(request, f'An error occurred: {str(e)}')
        return render(request, '404.html')
    
@staff_required
def update_product(request): 
    try:
        if request.method == 'POST':
            product_id = request.POST.get('txt_id') 
            product = get_object_or_404(Product, id=product_id)
            form = ProductUpdateForm(request.POST, instance=product)
            if form.is_valid():
                form.save()
                messages.success(request, 'Product Updated Successfully.')
                return redirect('/staff/product_list/')
            else:
                error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
                messages.warning(request, error_message)
                return redirect('/staff/product_list/')
        else:
            return redirect('/staff/product_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

@staff_required
def delete_product(request, id):
    try:
        product = get_object_or_404(Product, id=id)
        product.delete()
        messages.success(request, 'Product Deleted Successfully.')
        return redirect('/staff/product_list/')
    except Http404:
        return render(request, '404.html')

@staff_required
def purchase_list(request):
    try:
        stock_rec = Purchase.objects.select_related().order_by('-id')

        # Add pagination
        page = request.GET.get('page', 1)
        paginator = Paginator(stock_rec, 25)  # Show 10 purchases per page
        try:
            stock_rec = paginator.page(page)
        except PageNotAnInteger:
            stock_rec = paginator.page(1)
        except EmptyPage:
            stock_rec = paginator.page(paginator.num_pages)

        create_fm = PurchaseCreationForm()
        update_fm = PurchaseUpdateForm()
        context = {'stock_rec': stock_rec, 'create_fm': create_fm, 'update_fm': update_fm}
        return render(request, 'staff__purchase_list.html', context)
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')
    

@staff_required
def create_purchase(request):
    try:
        form = PurchaseCreationForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Record Added Successfully.....')
        else:
            error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
            messages.warning(request, error_message)
        return redirect('/staff/product_list/')
    except Exception as e:
        return render(request, '404.html')

# def update_purchase(request): 
#     if request.method == 'POST':
#         purchase_id=request.POST.get('txt_id') 
#         purchase = get_object_or_404(Purchase, id=purchase_id)
#         form = PurchaseUpdateForm(request.POST, instance=purchase)
#         if form.is_valid():
#             form.save()
#             messages.success(request, 'Stock Updated Successfully.')
#             return redirect('/staff/purchase_list/')
#         else:
#             messages.warning(request, 'Form Not Submitted: Something Wrong')
#             return redirect('/staff/purchase_list/')
#     else:
#         messages.warning(request, 'Form Not Submited : Something Missing')
#         return redirect('/staff/purchase_list/')
    
@staff_required
def delete_purchase(request, id):
    try:
        purchase = get_object_or_404(Purchase, id=id)

        if purchase.quantity == 0:
            purchase.delete()
            messages.success(request, 'Stock Deleted Successfully.')
        else:
            qty = purchase.quantity
            stock = purchase.product.available_stock  # Assuming 'product' is a ForeignKey in Purchase model

            if qty > stock:
                messages.info(request, 'Quantity should be less than Stock')
                return redirect('/staff/purchase_list/')

            update_qty = int(stock) - int(qty)

            # Fetch the related Product instance and update its available_stock
            product_instance = purchase.product
            product_instance.available_stock = update_qty
            product_instance.save()

            purchase.delete()
            messages.success(request, 'Stock Deleted Successfully.')

        return redirect('/staff/purchase_list/')
    except Http404:
        return render(request, '404.html')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')
 
 
 

@staff_required
def invoice_list(request):
    try:
        rec = Invoice.objects.select_related().order_by('-id')
        Filter = Invoice_List_Filter(request.GET, queryset=rec)
        invoice_rec = Filter.qs 

        # Add pagination
        page = request.GET.get('page', 1)
        paginator = Paginator(invoice_rec, 25)  # Show 10 invoices per page
        try:
            invoice_rec = paginator.page(page)
        except PageNotAnInteger:
            invoice_rec = paginator.page(1)
        except EmptyPage:
            invoice_rec = paginator.page(paginator.num_pages)

        create_invoice_fm = InvoiceCreationForm()
        update_invoice_fm = InvoiceUpdateForm()
        context = {
            'invoice_rec': invoice_rec,
            'create_invoice_fm': create_invoice_fm,
            'update_invoice_fm': update_invoice_fm,
            'filter': Filter
        }
        return render(request, 'staff__invoice_list.html', context)
    except Exception as e:
        return render(request, '404.html')
    

@staff_required
def create_invoice(request):
    try:
        form = InvoiceCreationForm(request.POST)
        if form.is_valid():
            saved_instance = form.save()
            saved_id = saved_instance.id
            return redirect(f'/staff/invoice_item_list/{saved_id}')
        else:
            error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
            messages.warning(request, error_message)
        return redirect('/staff/invoice_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html') 

@staff_required
def update_invoice(request): 
    try:
        if request.method == 'POST':
            invoice_id = request.POST.get('txt_id') 
            invoice = get_object_or_404(Invoice, id=invoice_id)
            form = InvoiceUpdateForm(request.POST, instance=invoice)
            if form.is_valid():
                form.save()
                messages.success(request, 'Invoice Updated Successfully.')
                return redirect('/staff/invoice_list/')
            else:
                error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
                messages.warning(request, error_message)
        else:
            messages.warning(request, 'Form Not Updated: Something Missing')
            return redirect('/staff/invoice_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')


@staff_required
def delete_invoice(request, id):
    try:
        invoice = get_object_or_404(Invoice, id=id)

        # Check if there are associated InvoiceItems
        if invoice.invoiceitem_set.exists():
            messages.warning(request, 'Invoice not deleted. Invoice is not empty.')
            return redirect('/staff/invoice_list/')

        # If no associated InvoiceItems, proceed with deletion
        with transaction.atomic():
            invoice.delete()
            messages.success(request, 'Invoice Deleted Successfully...')
            return redirect('/staff/invoice_list/')
    except Http404:
        return render(request, '404.html')

@staff_required
def add_invoice_in_account(request,id):
    try:
        invoice=get_object_or_404(Invoice, id=id)
        check_invoice_exist=Account.objects.filter(invoice_number=invoice.invoice_number).exists()
        
        if not invoice.grand_total:
            messages.warning(request, 'Amount not available.....') 
            return redirect('/staff/invoice_list/') 
        
        if check_invoice_exist:
            messages.warning(request, 'Invoice Exist.....') 
            return redirect('/staff/invoice_list')
        else: 
            acc_rec_count = Account.objects.filter(dealer=invoice.dealer).count()
            if acc_rec_count:
                balance_rec = Account.objects.filter(dealer=invoice.dealer).latest('id')
                running_balance = balance_rec.balance - int(invoice.grand_total) if acc_rec_count > 0 else int(invoice.grand_total)
            else:
                running_balance = int(invoice.grand_total)
            
            Account.objects.create(
                dealer=invoice.dealer,
                amount=int(invoice.grand_total),
                payment_mode='other',  # Add the appropriate payment mode
                is_credit=False,  # Adjust based on your logic
                balance=running_balance,
                invoice_number=invoice.invoice_number,
                created_by=request.user,  # Assuming you have a user associated with the request
            )
            Invoice.objects.filter(id=id).update(is_added_in_account=True)
            messages.success(request, 'Invoice Added in Account.....') 
            return redirect('/staff/invoice_list/') 
    except Exception as e:
            return render(request, '404.html')
        
@staff_required
def invoice_item_list(request, id):
    request.session['session_invoice_id']=id
    records = InvoiceItem.objects.filter(invoice=id)
    sale_rec = Invoice.objects.get(id=id) 
    form = InvoiceProductForm()
    if request.method == 'POST': 
        form = InvoiceProductForm(request.POST)
        if form.is_valid():
            product_id = form.cleaned_data['product'] 
            product_qty = form.cleaned_data['quantity'] 
            available_stock=Product.objects.get(id=product_id.id).available_stock
            if int(available_stock) < int(product_qty):
                messages.warning(request, f'Only {available_stock} Quantity Available')
                return redirect(f'/staff/invoice_item_list/{id}')
            else:
                fm = form.save(commit=False)
                fm.invoice = sale_rec  # Associate the form with the specific invoice
                fm.save()
                messages.success(request, 'Item Added Successfully')
                return redirect(f'/staff/invoice_item_list/{id}')
        else:
            messages.warning(request, 'Item Not Added')

    total_quantity = InvoiceItem.objects.filter(invoice=sale_rec).aggregate(total_quantity=Sum('quantity'))['total_quantity']
    total_gst_amount = InvoiceItem.objects.filter(invoice=sale_rec).aggregate(total_gst_amount=Sum('gst_amount'))['total_gst_amount']
    grand_total_amount = InvoiceItem.objects.filter(invoice=sale_rec).aggregate(total_total_amount=Sum('total_amount'))['total_total_amount']
    total_products = InvoiceItem.objects.filter(invoice=sale_rec).count()
 
    context = {'form': form,'id':id, 'invoice_rec': records,'total_quantity':total_quantity,'total_gst_amount':total_gst_amount,'grand_total_amount':grand_total_amount,'total_products':total_products,'invoice_details':sale_rec}
    return render(request, 'staff__manage_invoice_item.html', context)


@staff_required
def delete_invoice_item(request, id):
    try:
        invoice = get_object_or_404(InvoiceItem, id=id)
        invoice.delete()
        id=request.session.get('session_invoice_id')
        messages.success(request, 'Product Deleted Successfully.')
        return redirect(f'/staff/invoice_item_list/{id}')
    except Http404:
        return render(request, '404.html')


@staff_required
def get_product_details(request):
    product_code = request.GET.get('product_code', '')
    product_id = request.GET.get('productId', '')  # Use 'product_id' to match the parameter in your AJAX request
    product = None 

    if product_code:
        product = Product.objects.filter(product_code=product_code).first()
    elif product_id:
        product = Product.objects.filter(id=product_id).first()
    if product:
        data = {
            'product_name': product.id,
            'sale_amount': product.sale_amount,
            'gst': product.gst,
            'available_stock': product.available_stock,
            # Include other product details as needed
        }
        return JsonResponse(data)
    else:
        return JsonResponse({'error': 'Product not found'})
    

import openpyxl

@staff_required
def dealer_bulk_creation(request):
    if request.method == "POST":
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            try:
                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook.active 
                data_to_insert = []
                
                num_records_inserted=0
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    dealer_name = row[0]
                    business_name = row[1]
                    mobile_no = row[2]
                    email_id = row[3]
                    address = row[4]
                    state_name = row[5]  # Assuming the state name is provided in the Excel file
                    pin_code = row[6]
                    gst_number = row[7]

 
                    state_exist=State.objects.filter(state_name=state_name).exists()
                    if not state_exist:
                        messages.warning(request, f"{state_name} State Not Exist...!")
                        return redirect('/staff/dealer_list/')
                
                    num_records_inserted+=1
                    # Fetch the existing state or create a new one if it doesn't exist
                    state_obj, created = State.objects.get_or_create(state_name=state_name)
                    
                    dealer_obj = Dealer(
                        dealer_name=dealer_name,
                        business_name=business_name,
                        mobile_no=mobile_no,
                        email_id=email_id,
                        address=address,
                        state=state_obj,
                        pin_code=pin_code,
                        gst_number=gst_number
                    )
                    data_to_insert.append(dealer_obj)
                
                Dealer.objects.bulk_create(data_to_insert)
                messages.success(request, f'{num_records_inserted} records inserted successfully')
            except Exception as e:
                messages.error(request, f'Error occurred during import: {str(e)}')
        else:
            messages.error(request, 'No file selected.')
        return redirect('/staff/dealer_list/')


@staff_required
def product_bulk_creation(request):
    if request.method == "POST":
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            try:
                workbook = openpyxl.load_workbook(excel_file)
                worksheet = workbook.active 
                data_to_insert = []
                
                num_records_inserted = 0
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    product_code = row[0]
                    product_name = row[1]
                    description = row[2]
                    sale_amount = row[3]
                    hsn_sac = row[4]
                    gst = row[5]
                    available_stock = row[6]

                    # Check if product with the same code already exists
                    product_exist = Product.objects.filter(product_code=product_code).exists()
                    if product_exist:
                        messages.warning(request, f"Product with code {product_code} already exists.")
                        return redirect('/staff/product_list/')
                
                    num_records_inserted += 1

                    product_obj = Product(
                        product_code=product_code,
                        product_name=product_name,
                        description=description,
                        sale_amount=sale_amount,
                        hsn_sac=hsn_sac,
                        gst=gst,
                        available_stock=available_stock
                    )
                    data_to_insert.append(product_obj)
                
                Product.objects.bulk_create(data_to_insert)
                messages.success(request, f'{num_records_inserted} records inserted successfully')
            except Exception as e:
                messages.error(request, f'Error occurred during import: {str(e)}')
        else:
            messages.error(request, 'No file selected.')
        return redirect('/staff/product_list/')



@staff_required
def print_all_invoice_formate(request, id):
    try:
        invoice_details = Invoice.objects.get(id=id)
        item_rec = InvoiceItem.objects.filter(invoice=id)
         
        total_quantity = InvoiceItem.objects.filter(invoice=invoice_details).aggregate(total_quantity=Sum('quantity'))['total_quantity']
        total_gst_amount = InvoiceItem.objects.filter(invoice=invoice_details).aggregate(total_gst_amount=Sum('gst_amount'))['total_gst_amount']
        grand_total_amount = InvoiceItem.objects.filter(invoice=invoice_details).aggregate(total_total_amount=Sum('total_amount'))['total_total_amount']
        total_rate = item_rec.aggregate(total_rate=Sum('rate'))['total_rate']
        total_taxable_amount = item_rec.aggregate(total_taxable_amount=Sum('taxable_amount'))['total_taxable_amount']

        if not str(invoice_details.dealer.state) == "Gujarat":
            gst_type="IGST"
        else: 
            gst_type="CGST / SGST"
 

        context = {'invoice_details': invoice_details, 
                   'item_rec': item_rec,
                   'total_quantity':total_quantity,
                   'total_gst_amount':total_gst_amount,
                   'grand_total_amount':grand_total_amount,
                   'total_rate':total_rate,
                   'total_taxable_amount':total_taxable_amount,
                   'gst_type':gst_type,
                   }
        return render(request, 'staff__print_all_invoice_formate.html', context)
    except Invoice.DoesNotExist:
        return render(request, '404.html')
    
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage

@staff_required
def transaction_list(request):
    try: 
        rec = Account.objects.select_related().order_by('-id')
        Filter = Transaction_List_Filter(request.GET, queryset=rec)
        transaction_rec = Filter.qs 

        # Add pagination
        page = request.GET.get('page', 1)
        paginator = Paginator(transaction_rec, 25)  # Show 10 transactions per page
        try:
            transaction_rec = paginator.page(page)
        except PageNotAnInteger:
            transaction_rec = paginator.page(1)
        except EmptyPage:
            transaction_rec = paginator.page(paginator.num_pages)

        latest_transaction = Account.objects.order_by('-id').first()
        latest_transaction_id = latest_transaction.id if latest_transaction else None
        create_transaction_fm = CreateTransactionForm()
        update_transaction_fm = UpdateTransactionForm() 
        context = {
            'latest_transaction_id': latest_transaction_id,
            'transaction_rec': transaction_rec,
            'create_transaction_fm': create_transaction_fm,
            'update_transaction_fm': update_transaction_fm,
            'filter': Filter
        }
        return render(request, 'staff__transactions_list.html', context)
    except Exception as e:
        return render(request, '404.html')


@staff_required
def create_transaction(request):
    try:
        if request.method == 'POST':
            acc_rec_count = Account.objects.filter(dealer=request.POST.get('dealer')).count()
            if acc_rec_count:
                balance_rec = Account.objects.filter(dealer=request.POST.get('dealer')).latest('id')
            payment_status = request.POST.get('transaction-type-create')
   
            form = CreateTransactionForm(request.POST)
            if form.is_valid():
                fm = form.save(commit=False)
                if payment_status == 'credit':
                    fm.is_credit=True
                    fm.balance = balance_rec.balance +  int(request.POST.get('amount')) if acc_rec_count > 0 else int(request.POST.get('amount'))
                else:
                    fm.is_credit=False
                    fm.balance = balance_rec.balance -  int(request.POST.get('amount')) if acc_rec_count > 0 else int(request.POST.get('amount'))
                fm.created_by=request.user
                fm.save()
                messages.success(request, 'Transaction Created Successfully.....')
            else:
                error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
                messages.warning(request, error_message)
            return redirect('/staff/transaction_list/')
        else:
            messages.warning(request, 'Something is missing ....!')
            return redirect('/staff/transaction_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

    
@staff_required
def update_transaction(request): 
    try:
        if request.method == 'POST':
            transaction_id = request.POST.get('txt_id') 
            update_payment_status = request.POST.get('transaction-type-update')
 
            transaction = get_object_or_404(Account, id=transaction_id)
            form = UpdateTransactionForm(request.POST, instance=transaction)
            if form.is_valid():
                fm = form.save(commit=False)
                if update_payment_status == 'credit':
                    fm.is_credit=True
                else:
                    fm.is_credit=False 

                fm.modified_by=request.user
                fm.save()
                messages.success(request, 'Trasaction Updated Successfully.')
                return redirect('/staff/troubleshoot_transactions_for_balance/')
            else:
                error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
                messages.warning(request, error_message)
                return redirect('/staff/transaction_list/')
        else:
            return redirect('/staff/transaction_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

@staff_required
def delete_transaction(request, id):
    try:
        # trans_for_invoice = get_object_or_404(Account, id=id)
        # invoice = get_object_or_404(Invoice, invoice_number=trans_for_invoice.invoice_number)
        # Invoice.objects.filter(id=invoice.id).update(is_added_in_account=False)

        transaction = get_object_or_404(Account, id=id)
        transaction.delete()
        messages.success(request, 'Transaction Deleted Successfully.')
        return redirect('/staff/troubleshoot_transactions_for_balance/')
    except Http404:
        return render(request, '404.html')


@staff_required
def troubleshoot_transactions_for_balance(request):
    try:

        dealers = Dealer.objects.all()

        for dealer in dealers:
            # Filter previous entries for the same dealer
            previous_entries = Account.objects.filter(dealer=dealer).order_by('transaction_date')

            for account in previous_entries:
                # Calculate previous credited and debited amounts
                previous_credited_amount = previous_entries.filter(
                    transaction_date__lte=account.transaction_date, is_credit=True
                ).aggregate(Sum('amount'))['amount__sum'] or 0

                previous_debited_amount = previous_entries.filter(
                    transaction_date__lte=account.transaction_date, is_credit=False
                ).aggregate(Sum('amount'))['amount__sum'] or 0

                # Update the balance field
                account.balance = previous_credited_amount - previous_debited_amount
                account.save()

        return redirect('/staff/transaction_list/')
    except Exception as e:
        return render(request, '404.html')
    




@staff_required
def performa_invoice_list(request):
    try:
        rec = PerformaInvoice.objects.select_related().order_by('-id')
        Filter=Performa_Invoice_List_Filter(request.GET, queryset=rec)
        invoice_rec=Filter.qs 
        create_invoice_fm = PerformaInvoiceCreationForm()
        update_invoice_fm = PerformaInvoiceUpdateForm()
        context = {'invoice_rec': invoice_rec, 'create_invoice_fm': create_invoice_fm, 'update_invoice_fm': update_invoice_fm,'filter':Filter}
        return render(request, 'staff__performa_invoice_list.html', context)
    except Exception as e:
        return render(request, '404.html')

@staff_required
def create_performa_invoice(request):
    try:
        form = PerformaInvoiceCreationForm(request.POST)
        if form.is_valid():
            saved_instance = form.save()
            saved_id = saved_instance.id
            return redirect(f'/staff/performa_invoice_item_list/{saved_id}')
        else:
            error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
            messages.warning(request, error_message)
        return redirect('/staff/performa_invoice_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')

@staff_required
def update_performa_invoice(request): 
    try:
        if request.method == 'POST':
            performa_invoice_id = request.POST.get('txt_id') 
            performa_invoice = get_object_or_404(PerformaInvoice, id=performa_invoice_id)
            form = PerformaInvoiceUpdateForm(request.POST, instance=performa_invoice)
            if form.is_valid():
                form.save()
                messages.success(request, 'Invoice Updated Successfully.')
                return redirect('/staff/performa_invoice_list/')
            else:
                error_message = ', '.join([f"{field}: {error[0]}" for field, error in form.errors.items()])
                messages.warning(request, error_message)
        else:
            messages.warning(request, 'Form Not Updated: Something Missing')
            return redirect('/staff/invoice_list/')
    except Exception as e:
        # Log the exception if needed
        return render(request, '404.html')


@staff_required
def delete_performa_invoice(request, id):
    try:
        performa_invoice = get_object_or_404(PerformaInvoice, id=id)
        with transaction.atomic():
            performa_invoice.delete()
            messages.success(request, 'Performa Invoice Deleted Successfully...')
            return redirect('/staff/performa_invoice_list/')
    except Http404:
        return render(request, '404.html')



     
@staff_required
def performa_invoice_item_list(request, id):
    request.session['session_invoice_id']=id
    records = PerformaInvoiceItem.objects.filter(performa_invoice=id)
    sale_rec = PerformaInvoice.objects.get(id=id) 
    form = PerformaInvoiceProductForm()
    if request.method == 'POST': 
        form = PerformaInvoiceProductForm(request.POST)
        if form.is_valid(): 
            fm = form.save(commit=False)
            fm.performa_invoice = sale_rec  
            fm.save()
            messages.success(request, 'Item Added Successfully')
            return redirect(f'/staff/performa_invoice_item_list/{id}')
        else:
            messages.warning(request, 'Item Not Added')

    total_quantity = PerformaInvoiceItem.objects.filter(performa_invoice=sale_rec).aggregate(total_quantity=Sum('quantity'))['total_quantity']
    total_gst_amount = PerformaInvoiceItem.objects.filter(performa_invoice=sale_rec).aggregate(total_gst_amount=Sum('gst_amount'))['total_gst_amount']
    grand_total_amount = PerformaInvoiceItem.objects.filter(performa_invoice=sale_rec).aggregate(total_total_amount=Sum('total_amount'))['total_total_amount']
    total_products = PerformaInvoiceItem.objects.filter(performa_invoice=sale_rec).count()
 
    context = {'form': form,'id':id, 'invoice_rec': records,'total_quantity':total_quantity,'total_gst_amount':total_gst_amount,'grand_total_amount':grand_total_amount,'total_products':total_products,'invoice_details':sale_rec}
    return render(request, 'staff__performa_invoice_item.html', context)


@staff_required
def delete_performa_invoice_item(request, id):
    try:
        invoice = get_object_or_404(PerformaInvoiceItem, id=id)
        invoice.delete()
        id=request.session.get('session_invoice_id')
        messages.success(request, 'Product Deleted Successfully.')
        return redirect(f'/staff/performa_invoice_item_list/{id}')
    except Http404:
        return render(request, '404.html')


@staff_required
def print_performa_invoice(request, id):
    try:
        invoice_details = PerformaInvoice.objects.get(id=id)
        item_rec = PerformaInvoiceItem.objects.filter(performa_invoice=id)
         
        total_quantity = PerformaInvoiceItem.objects.filter(performa_invoice=invoice_details).aggregate(total_quantity=Sum('quantity'))['total_quantity']
        total_gst_amount = PerformaInvoiceItem.objects.filter(performa_invoice=invoice_details).aggregate(total_gst_amount=Sum('gst_amount'))['total_gst_amount']
        grand_total_amount = PerformaInvoiceItem.objects.filter(performa_invoice=invoice_details).aggregate(total_total_amount=Sum('total_amount'))['total_total_amount']
        total_rate = item_rec.aggregate(total_rate=Sum('rate'))['total_rate']
        total_taxable_amount = item_rec.aggregate(total_taxable_amount=Sum('taxable_amount'))['total_taxable_amount']

        if not str(invoice_details.state) == "Gujarat":
            gst_type="IGST"
        else: 
            gst_type="CGST / SGST"
 

        context = {'invoice_details': invoice_details, 
                   'item_rec': item_rec,
                   'total_quantity':total_quantity,
                   'total_gst_amount':total_gst_amount,
                   'grand_total_amount':grand_total_amount,
                   'total_rate':total_rate,
                   'total_taxable_amount':total_taxable_amount,
                   'gst_type':gst_type,
                   }
        return render(request, 'staff__print_performa_invoice.html', context)
    except Invoice.DoesNotExist:
        return render(request, '404.html')
    

 
@staff_required
def print_transactions_receipt(request,id):
    try: 
        data=Account.objects.get(id=id)
        context={'data':data}
        return render(request, 'staff__print_payment_received_slip.html', context)
    except Exception as e:
        return render(request, '404.html')
